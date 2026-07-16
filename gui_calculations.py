import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from payroll_calculator import PayrollCalculator
import logging

class CalculationsTab:
    def __init__(self, parent, db, calendar, app):
        self.parent = parent
        self.db = db
        self.calendar = calendar
        self.app = app
        self.clients = []
        self.calculator = PayrollCalculator(calendar)
        self.current_employees = []
        self.editable_entries = {}
        
        self.create_widgets()
        logging.info("Вкладка 'Расчеты' создана")

    def create_widgets(self):
        # Верхняя панель выбора
        top_frame = ttk.LabelFrame(self.parent, text="Параметры расчета", padding=10)
        top_frame.pack(fill='x', padx=5, pady=5)
        
        # Первая строка
        row1 = ttk.Frame(top_frame)
        row1.pack(fill='x', pady=2)
        
        ttk.Label(row1, text="🏢 Клиент:").pack(side='left', padx=5)
        self.calc_client_cb = ttk.Combobox(row1, state='readonly', width=30)
        self.calc_client_cb.pack(side='left', padx=5)
        self.calc_client_cb.bind('<<ComboboxSelected>>', self.on_client_changed)
        
        ttk.Label(row1, text="📅 Месяц:").pack(side='left', padx=5)
        self.month_var = tk.StringVar()
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                 "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        self.month_cb = ttk.Combobox(row1, textvariable=self.month_var, 
                                     state='readonly', width=15, values=months)
        self.month_cb.current(datetime.now().month - 1)
        self.month_cb.pack(side='left', padx=5)
        self.month_cb.bind('<<ComboboxSelected>>', self.update_month_info)
        
        ttk.Label(row1, text="📆 Год:").pack(side='left', padx=5)
        self.year_var = tk.StringVar()
        years = [str(y) for y in range(2024, 2031)]
        self.year_cb = ttk.Combobox(row1, textvariable=self.year_var, 
                                   state='readonly', width=8, values=years)
        self.year_cb.current(years.index(str(datetime.now().year)))
        self.year_cb.pack(side='left', padx=5)
        self.year_cb.bind('<<ComboboxSelected>>', self.update_month_info)
        
        # Вторая строка - тип выплаты и кнопка
        row2 = ttk.Frame(top_frame)
        row2.pack(fill='x', pady=5)
        
        self.payment_type = tk.StringVar(value="salary")
        ttk.Radiobutton(row2, text="💵 Аванс (до 15 числа)", 
                       variable=self.payment_type, 
                       value="advance",
                       command=self.on_payment_type_changed).pack(side='left', padx=10)
        ttk.Radiobutton(row2, text="💰 Зарплата (за месяц)", 
                       variable=self.payment_type, 
                       value="salary",
                       command=self.on_payment_type_changed).pack(side='left', padx=10)
        
        ttk.Button(row2, text="🧮 Рассчитать", 
                  command=self.calculate_payroll).pack(side='right', padx=20)
        ttk.Button(row2, text="📂 Загрузить сохраненное", 
                  command=self.load_saved).pack(side='right', padx=5)
        
        # Информация о месяце
        self.month_info_label = ttk.Label(top_frame, text="", foreground='blue')
        self.month_info_label.pack(pady=5)
        
        # Основная панель с таблицей и детализацией
        main_panel = ttk.Frame(self.parent)
        main_panel.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Левая часть - таблица с сотрудниками
        table_frame = ttk.LabelFrame(main_panel, text="Ведомость начисления", padding=10)
        table_frame.pack(side='left', fill='both', expand=True)
        
        # Создаем Treeview с прокрутками
        columns = ('name', 'salary', 'norm_days', 'worked_days', 'accrued', 'ndfl', 'to_pay')
        self.result_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=12)
        
        self.result_tree.heading('name', text='Сотрудник')
        self.result_tree.heading('salary', text='Оклад (руб.)')
        self.result_tree.heading('norm_days', text='Норма дней')
        self.result_tree.heading('worked_days', text='Отраб. дней')
        self.result_tree.heading('accrued', text='Начислено')
        self.result_tree.heading('ndfl', text='НДФЛ')
        self.result_tree.heading('to_pay', text='К выплате')
        
        self.result_tree.column('name', width=180)
        self.result_tree.column('salary', width=100, anchor='e')
        self.result_tree.column('norm_days', width=90, anchor='center')
        self.result_tree.column('worked_days', width=90, anchor='center')
        self.result_tree.column('accrued', width=110, anchor='e')
        self.result_tree.column('ndfl', width=100, anchor='e')
        self.result_tree.column('to_pay', width=110, anchor='e')
        
        # Добавляем прокрутки
        tree_scroll = ttk.Scrollbar(table_frame, orient='vertical', command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=tree_scroll.set)
        
        self.result_tree.pack(side='left', fill='both', expand=True)
        tree_scroll.pack(side='right', fill='y')
        
        # Привязываем двойной клик для редактирования
        self.result_tree.bind('<Double-Button-1>', self.on_tree_double_click)
        
        # Итоговые суммы
        totals_frame = ttk.Frame(table_frame)
        totals_frame.pack(fill='x', pady=5)
        
        self.total_label = ttk.Label(totals_frame, text="", font=('Arial', 10, 'bold'))
        self.total_label.pack(side='left', padx=10)
        
        # Правая часть - детализация НДФЛ
        detail_frame = ttk.LabelFrame(main_panel, text="Детализация НДФЛ", padding=10)
        detail_frame.pack(side='right', fill='both', padx=(10, 0))
        
        self.detail_tree = ttk.Treeview(detail_frame, 
                                       columns=('name', 'ndfl_advance', 'ndfl_second', 'ndfl_total'), 
                                       show='headings', height=12)
        
        self.detail_tree.heading('name', text='Сотрудник')
        self.detail_tree.heading('ndfl_advance', text='НДФЛ с аванса')
        self.detail_tree.heading('ndfl_second', text='НДФЛ со 2-й пол.')
        self.detail_tree.heading('ndfl_total', text='НДФЛ всего')
        
        self.detail_tree.column('name', width=150)
        self.detail_tree.column('ndfl_advance', width=120, anchor='e')
        self.detail_tree.column('ndfl_second', width=120, anchor='e')
        self.detail_tree.column('ndfl_total', width=120, anchor='e')
        
        self.detail_tree.pack(fill='both', expand=True)
        
        self.detail_total_label = ttk.Label(detail_frame, text="", font=('Arial', 9, 'bold'))
        self.detail_total_label.pack(pady=5)
        
        # Кнопки действий
        btn_frame = ttk.Frame(self.parent)
        btn_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(btn_frame, text="📋 Копировать в буфер", 
                  command=self.copy_to_clipboard).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📝 Отправить в сообщения", 
                  command=self.send_to_messages).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📊 Экспорт в Excel", 
                  command=self.export_to_excel).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="💾 Сохранить в БД", 
                  command=self.save_to_db).pack(side='left', padx=5)

    def refresh_clients_list(self, clients):
        """Обновление списка клиентов"""
        self.clients = clients
        client_names = [f"{c[1]} (ИНН: {c[2]})" for c in clients]
        self.calc_client_cb['values'] = client_names
        if client_names and not self.calc_client_cb.get():
            self.calc_client_cb.current(0)
            self.update_month_info()

    def on_client_changed(self, event=None):
        """Обработчик смены клиента"""
        self.update_month_info()
        self.clear_results()

    def update_month_info(self, event=None):
        """Обновление информации о выбранном месяце"""
        try:
            month_name = self.month_var.get()
            year = int(self.year_var.get())
            months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                     "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
            
            if month_name in months:
                month = months.index(month_name) + 1
                info = self.calendar.get_month_info(year, month)
                self.month_info_label.config(
                    text=f"📊 {info['name']} {year}: {info['workdays']} рабочих дней, "
                         f"{info['weekends']} выходных, {info['total_days']} всего"
                )
        except Exception as e:
            logging.error(f"Ошибка обновления информации о месяце: {e}")

    def on_payment_type_changed(self):
        """Обработчик смены типа выплаты"""
        if self.current_employees:
            self.recalculate_all()

    def get_selected_client_id(self):
        """Получение ID выбранного клиента"""
        selected = self.calc_client_cb.get()
        if not selected:
            return None
        
        for c in self.clients:
            if f"{c[1]} (ИНН: {c[2]})" == selected:
                return c[0]
        return None

    def get_selected_year_month(self):
        """Получение выбранного года и месяца"""
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                 "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        
        month_name = self.month_var.get()
        if month_name not in months:
            return None, None
        
        month = months.index(month_name) + 1
        
        try:
            year = int(self.year_var.get())
        except ValueError:
            return None, None
        
        return year, month

    def calculate_payroll(self):
        """Расчет зарплаты"""
        client_id = self.get_selected_client_id()
        if not client_id:
            messagebox.showwarning("⚠️ Внимание", "Выберите клиента!")
            return
        
        year, month = self.get_selected_year_month()
        if not year:
            messagebox.showerror("❌ Ошибка", "Выберите месяц и год!")
            return
        
        self.current_employees = self.db.get_active_employees_for_month(client_id, year, month)
        
        if not self.current_employees:
            self.clear_results()
            messagebox.showinfo("ℹ️ Информация", "Нет сотрудников для расчета за этот месяц.")
            return
        
        self.recalculate_all()
        
        payment_type = "аванса" if self.payment_type.get() == "advance" else "зарплаты"
        logging.info(f"Выполнен расчет {payment_type} для клиента ID={client_id} за {month}.{year}")

    def recalculate_all(self):
        """Пересчет всех сотрудников"""
        year, month = self.get_selected_year_month()
        if not year:
            return
        
        payment_type = self.payment_type.get()
        
        # Собираем ручные вводы дней
        worked_days_dict = {}
        for emp in self.current_employees:
            emp_id = emp[0]
            if emp_id in self.editable_entries:
                try:
                    days = int(self.editable_entries[emp_id].get())
                    if days >= 0:
                        worked_days_dict[emp_id] = days
                except ValueError:
                    pass
        
        result = self.calculator.calculate_all(
            self.current_employees, year, month, payment_type, worked_days_dict
        )
        
        self.update_results_table(result)
        self.update_detail_table(result)

    def update_results_table(self, result):
        """Обновление основной таблицы результатов"""
        self.result_tree.delete(*self.result_tree.get_children())
        self.editable_entries.clear()
        
        for emp_result in result['results']:
            calc = emp_result['calculation']
            emp_id = emp_result['id']
            emp_name = emp_result['name']
            
            # Ищем оклад
            salary = 0
            for emp in self.current_employees:
                if emp[0] == emp_id:
                    salary = emp[2]
                    break
            
            total_workdays = calc.get('total_workdays', 0)
            worked_days = calc.get('worked_days', 0)
            accrued = calc.get('full_salary', calc.get('amount_before_tax', 0))
            ndfl = calc.get('ndfl_full', calc.get('ndfl', 0))
            to_pay = calc.get('to_pay', 0)
            
            values = (
                emp_name,
                f"{salary:,}",
                total_workdays,
                worked_days,
                f"{accrued:,}",
                f"{ndfl:,}",
                f"{to_pay:,}"
            )
            
            item_id = self.result_tree.insert('', 'end', values=values, tags=(str(emp_id),))
            
            # Создаем редактируемую ячейку
            self.create_editable_cell(item_id, emp_id, worked_days, total_workdays)
        
        # Обновляем итоги
        total_to_pay = result['total_to_pay']
        total_ndfl = result['total_ndfl']
        
        self.total_label.config(
            text=f"💰 Итого к выплате: {total_to_pay:,} руб.  |  💸 НДФЛ удержан: {total_ndfl:,} руб."
        )

    def create_editable_cell(self, item_id, emp_id, worked_days, total_workdays):
        """Создание редактируемой ячейки для отработанных дней"""
        # Получаем координаты ячейки
        bbox = self.result_tree.bbox(item_id, column='worked_days')
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Создаем Entry поверх ячейки
        entry = tk.Entry(self.result_tree, justify='center', font=('Arial', 9))
        entry.insert(0, str(worked_days))
        entry.place(x=x, y=y, width=width, height=height)
        
        self.editable_entries[emp_id] = entry
        
        # Привязываем события
        entry.bind('<Return>', lambda e, eid=emp_id: self.on_days_changed(eid))
        entry.bind('<FocusOut>', lambda e, eid=emp_id: self.on_days_changed(eid))
        entry.bind('<Up>', lambda e: self.focus_previous_entry(e.widget))
        entry.bind('<Down>', lambda e: self.focus_next_entry(e.widget))

    def on_days_changed(self, emp_id):
        """Обработчик изменения количества дней"""
        try:
            entry = self.editable_entries.get(emp_id)
            if not entry:
                return
            
            days_str = entry.get().strip()
            if not days_str:
                return
            
            days = int(days_str)
            
            year, month = self.get_selected_year_month()
            if not year:
                return
            
            total_workdays = self.calendar.get_workdays_in_month(year, month)
            
            if days > total_workdays:
                entry.config(foreground='red')
                self.app.set_status(f"⚠️ Количество дней не может превышать норму ({total_workdays})!")
                return
            elif days < 0:
                entry.config(foreground='red')
                self.app.set_status("⚠️ Количество дней не может быть отрицательным!")
                return
            else:
                entry.config(foreground='black')
            
            self.recalculate_single(emp_id, days)
            
        except ValueError:
            entry.config(foreground='red')
            self.app.set_status("⚠️ Введите корректное число дней!")
        except Exception as e:
            logging.error(f"Ошибка при изменении дней: {e}")

    def recalculate_single(self, emp_id, worked_days):
        """Пересчет одного сотрудника"""
        year, month = self.get_selected_year_month()
        if not year:
            return
        
        payment_type = self.payment_type.get()
        
        emp_data = None
        for emp in self.current_employees:
            if emp[0] == emp_id:
                emp_data = emp
                break
        
        if not emp_data:
            return
        
        calc = self.calculator.calculate_single(emp_data, year, month, worked_days, payment_type)
        
        # Обновляем строку в таблице
        for item_id in self.result_tree.get_children():
            if str(emp_id) in self.result_tree.item(item_id)['tags']:
                salary = emp_data[2]
                accrued = calc.get('full_salary', calc.get('amount_before_tax', 0))
                ndfl = calc.get('ndfl_full', calc.get('ndfl', 0))
                to_pay = calc.get('to_pay', 0)
                
                self.result_tree.item(item_id, values=(
                    emp_data[1],
                    f"{salary:,}",
                    calc.get('total_workdays', 0),
                    worked_days,
                    f"{accrued:,}",
                    f"{ndfl:,}",
                    f"{to_pay:,}"
                ))
                break
        
        self.recalculate_totals()
        self.recalculate_detail()

    def recalculate_totals(self):
        """Пересчет итоговых сумм"""
        total_to_pay = 0
        total_ndfl = 0
        
        for item_id in self.result_tree.get_children():
            values = self.result_tree.item(item_id)['values']
            try:
                ndfl_str = values[5].replace(',', '')
                to_pay_str = values[6].replace(',', '')
                total_ndfl += int(ndfl_str)
                total_to_pay += int(to_pay_str)
            except (ValueError, IndexError):
                pass
        
        self.total_label.config(
            text=f"💰 Итого к выплате: {total_to_pay:,} руб.  |  💸 НДФЛ удержан: {total_ndfl:,} руб."
        )

    def update_detail_table(self, result):
        """Обновление таблицы детализации НДФЛ"""
        self.detail_tree.delete(*self.detail_tree.get_children())
        
        if self.payment_type.get() == 'advance':
            self.detail_total_label.config(text="Для аванса НДФЛ удерживается единой суммой")
            return
        
        total_ndfl_advance = 0
        total_ndfl_second = 0
        total_ndfl = 0
        
        for emp_result in result['results']:
            calc = emp_result['calculation']
            
            ndfl_advance = calc.get('ndfl_advance', 0)
            ndfl_second = calc.get('ndfl_second_half', 0)
            ndfl_total = calc.get('ndfl_full', 0)
            
            self.detail_tree.insert('', 'end', values=(
                emp_result['name'],
                f"{ndfl_advance:,}",
                f"{ndfl_second:,}",
                f"{ndfl_total:,}"
            ))
            
            total_ndfl_advance += ndfl_advance
            total_ndfl_second += ndfl_second
            total_ndfl += ndfl_total
        
        self.detail_total_label.config(
            text=f"Итого: с аванса {total_ndfl_advance:,} руб. | "
                 f"со 2-й пол. {total_ndfl_second:,} руб. | "
                 f"всего {total_ndfl:,} руб."
        )

    def recalculate_detail(self):
        """Пересчет детализации НДФЛ"""
        year, month = self.get_selected_year_month()
        if not year:
            return
        
        payment_type = self.payment_type.get()
        if payment_type == 'advance':
            return
        
        worked_days_dict = {}
        for emp_id, entry in self.editable_entries.items():
            try:
                days = int(entry.get())
                if days >= 0:
                    worked_days_dict[emp_id] = days
            except ValueError:
                pass
        
        result = self.calculator.calculate_all(
            self.current_employees, year, month, payment_type, worked_days_dict
        )
        
        self.update_detail_table(result)

    def clear_results(self):
        """Очистка результатов"""
        self.result_tree.delete(*self.result_tree.get_children())
        self.detail_tree.delete(*self.detail_tree.get_children())
        self.editable_entries.clear()
        self.current_employees = []
        self.total_label.config(text="")
        self.detail_total_label.config(text="")

    def focus_previous_entry(self, current_entry):
        """Перемещение фокуса на предыдущий Entry"""
        entries = list(self.editable_entries.values())
        if current_entry in entries:
            idx = entries.index(current_entry)
            if idx > 0:
                entries[idx - 1].focus()

    def focus_next_entry(self, current_entry):
        """Перемещение фокуса на следующий Entry"""
        entries = list(self.editable_entries.values())
        if current_entry in entries:
            idx = entries.index(current_entry)
            if idx < len(entries) - 1:
                entries[idx + 1].focus()

    def on_tree_double_click(self, event):
        """Обработчик двойного клика по таблице"""
        region = self.result_tree.identify_region(event.x, event.y)
        if region == 'cell':
            column = self.result_tree.identify_column(event.x)
            if column == '#4':  # Колонка worked_days
                item_id = self.result_tree.identify_row(event.y)
                if item_id and item_id in self.result_tree.get_children():
                    for emp_id, entry in self.editable_entries.items():
                        if str(emp_id) in self.result_tree.item(item_id)['tags']:
                            entry.focus()
                            entry.select_range(0, tk.END)
                            break

    def save_to_db(self):
        """Сохранение расчета в базу данных"""
        if not self.current_employees:
            messagebox.showwarning("⚠️ Внимание", "Нет данных для сохранения!")
            return
        
        client_id = self.get_selected_client_id()
        year, month = self.get_selected_year_month()
        
        if not client_id or not year:
            messagebox.showerror("❌ Ошибка", "Выберите клиента, месяц и год!")
            return
        
        payment_type = self.payment_type.get()
        saved_count = 0
        
        for emp in self.current_employees:
            emp_id = emp[0]
            
            # Получаем отработанные дни
            worked_days = None
            if emp_id in self.editable_entries:
                try:
                    worked_days = int(self.editable_entries[emp_id].get())
                except ValueError:
                    worked_days = 0
            else:
                # Если не редактировали - берем из таблицы
                for item_id in self.result_tree.get_children():
                    if str(emp_id) in self.result_tree.item(item_id)['tags']:
                        values = self.result_tree.item(item_id)['values']
                        worked_days = int(values[3])
                        break
            
            if worked_days is None:
                continue
            
            # Выполняем расчет
            calc = self.calculator.calculate_single(emp, year, month, worked_days, payment_type)
            
            # Сохраняем в БД
            total_workdays = calc.get('total_workdays', 0)
            accrued = calc.get('full_salary', calc.get('amount_before_tax', 0))
            ndfl = calc.get('ndfl_full', calc.get('ndfl', 0))
            ndfl_advance = calc.get('ndfl_advance', 0)
            ndfl_second = calc.get('ndfl_second_half', 0)
            advance_paid = calc.get('advance_paid', 0)
            to_pay = calc.get('to_pay', 0)
            
            success = self.db.save_calculation(
                client_id, emp_id, year, month, payment_type,
                worked_days, total_workdays, accrued, ndfl,
                ndfl_advance, ndfl_second, advance_paid, to_pay
            )
            
            if success:
                saved_count += 1
        
        if saved_count > 0:
            self.app.set_status(f"✅ Сохранено {saved_count} расчетов")
            messagebox.showinfo("💾 Успех", f"Сохранено {saved_count} расчетов в базу данных!")
        else:
            messagebox.showerror("❌ Ошибка", "Не удалось сохранить расчеты!")

    def load_saved(self):
        """Загрузка сохраненных расчетов"""
        client_id = self.get_selected_client_id()
        year, month = self.get_selected_year_month()
        
        if not client_id or not year:
            messagebox.showwarning("⚠️ Внимание", "Выберите клиента, месяц и год!")
            return
        
        # Получаем сохраненные расчеты
        saved = self.db.get_saved_calculations(client_id, year, month)
        
        if not saved:
            messagebox.showinfo("ℹ️ Информация", "Нет сохраненных расчетов за этот период.")
            return
        
        # Определяем тип выплаты по первому сохраненному расчету
        payment_type = saved[0][2]  # payment_type в позиции 2
        
        # Устанавливаем тип выплаты
        self.payment_type.set(payment_type)
        
        # Получаем сотрудников
        self.current_employees = self.db.get_active_employees_for_month(client_id, year, month)
        
        if not self.current_employees:
            messagebox.showinfo("ℹ️ Информация", "Нет активных сотрудников для отображения.")
            return
        
        # Отображаем сохраненные данные
        self.result_tree.delete(*self.result_tree.get_children())
        self.editable_entries.clear()
        
        total_to_pay = 0
        total_ndfl = 0
        
        for calc_data in saved:
            # Распаковываем данные
            calc_id, emp_name, pay_type, worked_days, total_workdays, \
            accrued, ndfl, ndfl_advance, ndfl_second, advance_paid, to_pay, created_at = calc_data
            
            # Находим сотрудника
            emp_id = None
            salary = 0
            for emp in self.current_employees:
                if emp[1] == emp_name:
                    emp_id = emp[0]
                    salary = emp[2]
                    break
            
            if not emp_id:
                continue
            
            values = (
                emp_name,
                f"{salary:,}",
                total_workdays,
                worked_days,
                f"{accrued:,}",
                f"{ndfl:,}",
                f"{to_pay:,}"
            )
            
            item_id = self.result_tree.insert('', 'end', values=values, tags=(str(emp_id),))
            self.create_editable_cell(item_id, emp_id, worked_days, total_workdays)
            
            total_ndfl += int(ndfl)
            total_to_pay += int(to_pay)
        
        self.total_label.config(
            text=f"💰 Итого к выплате: {total_to_pay:,} руб.  |  💸 НДФЛ удержан: {total_ndfl:,} руб."
        )
        
        # Обновляем детализацию
        self.recalculate_detail()
        
        self.app.set_status(f"📂 Загружены сохраненные расчеты за {month}.{year}")

    def copy_to_clipboard(self):
        """Копирование результата в буфер обмена"""
        text = self.generate_report_text()
        if text:
            self.parent.clipboard_clear()
            self.parent.clipboard_append(text)
            messagebox.showinfo("✅ Успех", "Результат скопирован в буфер обмена!")
        else:
            messagebox.showwarning("⚠️ Внимание", "Нет данных для копирования!")

    def send_to_messages(self):
        """Отправка результата на вкладку Сообщения"""
        text = self.generate_report_text()
        if text:
            self.app.messages_tab.add_message(text)
            self.app.notebook.select(3)
            messagebox.showinfo("✅ Успех", "Результат отправлен на вкладку Сообщения!")
        else:
            messagebox.showwarning("⚠️ Внимание", "Нет данных для отправки!")

    def generate_report_text(self):
        """Генерация текстового отчета"""
        if not self.result_tree.get_children():
            return ""
        
        year, month = self.get_selected_year_month()
        month_name = self.month_var.get()
        payment_type = "Аванс" if self.payment_type.get() == "advance" else "Зарплата"
        
        lines = []
        lines.append(f"{'='*60}")
        lines.append(f"{payment_type.upper()} за {month_name} {year}")
        lines.append(f"{'='*60}")
        lines.append("")
        
        for item_id in self.result_tree.get_children():
            values = self.result_tree.item(item_id)['values']
            lines.append(f"👤 {values[0]}:")
            lines.append(f"   Оклад: {values[1]} руб.")
            lines.append(f"   Отработано дней: {values[3]} из {values[2]}")
            lines.append(f"   Начислено: {values[4]} руб.")
            lines.append(f"   НДФЛ: {values[5]} руб.")
            lines.append(f"   К выплате: {values[6]} руб.")
            lines.append("")
        
        lines.append(f"{'='*60}")
        lines.append(self.total_label.cget('text'))
        lines.append(f"{'='*60}")
        
        return "\n".join(lines)

    def export_to_excel(self):
        """Экспорт результатов в Excel"""
        try:
            import openpyxl
            
            if not self.result_tree.get_children():
                messagebox.showwarning("⚠️ Внимание", "Нет данных для экспорта!")
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Расчет зарплаты"
            
            # Заголовки
            headers = ['Сотрудник', 'Оклад', 'Норма дней', 'Отработано дней', 
                      'Начислено', 'НДФЛ', 'К выплате']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            for i, item_id in enumerate(self.result_tree.get_children(), 2):
                values = self.result_tree.item(item_id)['values']
                for j, val in enumerate(values, 1):
                    if j > 1:
                        try:
                            ws.cell(row=i, column=j, value=int(val.replace(',', '')))
                        except ValueError:
                            ws.cell(row=i, column=j, value=val)
                    else:
                        ws.cell(row=i, column=j, value=val)
            
            filename = f"Расчет_зарплаты_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("✅ Успех", f"Данные экспортированы в файл:\n{filename}")
            logging.info(f"Данные экспортированы в {filename}")
            
        except ImportError:
            messagebox.showerror("❌ Ошибка", 
                               "Для экспорта в Excel необходимо установить библиотеку openpyxl\n"
                               "Выполните команду: pip install openpyxl")
